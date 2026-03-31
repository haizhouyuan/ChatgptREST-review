#!/usr/bin/env python3
"""gen_review_sheet.py — 生成人工评分 Excel 表.

读取 planning_qa_scored.jsonl，生成按批次分组的 Excel 评分表，
预填机器评分供参考，留出人工评分空列。

用法:
    python gen_review_sheet.py                      # 全量生成
    python gen_review_sheet.py --batch-size 20      # 每批20条
    python gen_review_sheet.py --domain 两轮车车身业务  # 指定域

输出: review_batch_001.xlsx, review_batch_002.xlsx, ...

依赖: openpyxl (pip install openpyxl)
"""
from __future__ import annotations

import argparse
import json
import math
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

INPUT_DEFAULT = Path("/vol1/1000/projects/ChatgptREST/scripts/evomap_qa/planning_qa_scored.jsonl")
OUTPUT_DIR = Path("/vol1/1000/projects/ChatgptREST/scripts/evomap_qa/review_sheets")

# Styling
HEADER_FILL = PatternFill("solid", fgColor="2F5496") if HAS_OPENPYXL else None
MACHINE_FILL = PatternFill("solid", fgColor="D9E2F3") if HAS_OPENPYXL else None
HUMAN_FILL = PatternFill("solid", fgColor="FFF2CC") if HAS_OPENPYXL else None
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10) if HAS_OPENPYXL else None
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
) if HAS_OPENPYXL else None


def _truncate(text: str, max_len: int) -> str:
    if len(text) <= max_len:
        return text
    return text[:max_len-1] + "…"


def generate_xlsx(records: list[dict], output_path: Path, batch_num: int) -> None:
    """Generate a single Excel review sheet."""
    if not HAS_OPENPYXL:
        # Fallback to markdown
        generate_markdown(records, output_path.with_suffix(".md"), batch_num)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = f"评分批次_{batch_num:03d}"

    # Headers
    headers = [
        ("#", 5),
        ("QA ID", 14),
        ("域", 12),
        ("来源类型", 10),
        ("问题摘要", 40),
        ("答案摘要", 60),
        ("来源文件", 30),
        # Machine scores (info section)
        ("机器综合分", 10),
        ("路由判定", 12),
        # Human scoring columns (to be filled)
        ("清晰度\n(1-5)", 10),
        ("可行性\n(1-5)", 10),
        ("证据\n(1-5)", 10),
        ("风险\n(1-5)", 10),
        ("对齐\n(1-5)", 10),
        ("完整度\n(1-5)", 10),
        ("总分\n(1-5)", 10),
        ("评语", 30),
    ]

    for col_idx, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Set row height for header
    ws.row_dimensions[1].height = 35

    # Data rows
    for row_idx, rec in enumerate(records, 2):
        rubric = rec.get("rubric_auto", {})
        avg_rubric = sum(rubric.values()) / max(len(rubric), 1)
        # Convert 0-1 to 1-5 scale for display
        machine_score_display = round(1 + avg_rubric * 4, 1)

        row_data = [
            row_idx - 1,
            rec.get("qa_id", ""),
            rec.get("domain", ""),
            rec.get("source_type", ""),
            _truncate(rec.get("question", ""), 200),
            _truncate(rec.get("answer_summary", ""), 400),
            rec.get("source_file", ""),
            machine_score_display,
            rec.get("route_auto", ""),
            # Human scoring columns - leave empty
            None, None, None, None, None, None, None, "",
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER

            # Color machine score columns
            if col_idx in (8, 9):
                cell.fill = MACHINE_FILL
            # Color human score columns
            elif col_idx >= 10:
                cell.fill = HUMAN_FILL

        ws.row_dimensions[row_idx].height = 60

    # Add scoring guide sheet
    guide = wb.create_sheet("评分标尺")
    guide_data = [
        ["分值", "含义", "标准"],
        [5, "优秀", "可直接作为KB标准件，无需修改"],
        [4, "良好", "核心正确，小修即可使用"],
        [3, "及格", "方向正确但有明显缺漏，需补充"],
        [2, "不及格", "有重大错误或关键遗漏"],
        [1, "无效", "答非所问或完全错误"],
    ]
    for r, row in enumerate(guide_data, 1):
        for c, val in enumerate(row, 1):
            cell = guide.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = Font(bold=True)
            guide.column_dimensions[get_column_letter(c)].width = 40

    # Add dimension explanation sheet
    dims = wb.create_sheet("评分维度说明")
    dims_data = [
        ["维度", "英文", "评分要点"],
        ["清晰度", "clarity", "问题表述是否清楚？答案结构是否清晰、有逻辑？"],
        ["可行性", "feasibility", "方案/建议是否可落地执行？条件和约束是否合理？"],
        ["证据", "evidence", "结论是否有数据/文献/标准支撑？引用是否可核验？"],
        ["风险", "risk", "是否识别了主要风险？是否给出缓解措施？"],
        ["对齐", "alignment", "答案是否回答了问题？没有跑题或遗漏核心要求？"],
        ["完整度", "completeness", "覆盖面是否充分？有无重要方面被遗漏？"],
    ]
    for r, row in enumerate(dims_data, 1):
        for c, val in enumerate(row, 1):
            cell = dims.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = Font(bold=True)
            dims.column_dimensions[get_column_letter(c)].width = [12, 14, 50][c-1]

    wb.save(output_path)
    print(f"  Generated: {output_path.name} ({len(records)} records)")


def generate_markdown(records: list[dict], output_path: Path, batch_num: int) -> None:
    """Fallback: Generate markdown review table."""
    lines = [
        f"# 人工评分表 批次 {batch_num:03d}",
        f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "",
        "## 评分标尺: 1=无效, 2=不及格, 3=及格, 4=良好, 5=优秀",
        "",
        "| # | QA ID | 域 | 问题摘要 | 答案摘要 | 机器分 | 清晰度 | 可行性 | 证据 | 风险 | 对齐 | 完整度 | 总分 | 评语 |",
        "|---|---|---|---|---|---|---|---|---|---|---|---|---|---|",
    ]

    for i, rec in enumerate(records, 1):
        rubric = rec.get("rubric_auto", {})
        avg = sum(rubric.values()) / max(len(rubric), 1)
        ms = round(1 + avg * 4, 1)
        q = _truncate(rec.get("question", ""), 50).replace("|", "\\|").replace("\n", " ")
        a = _truncate(rec.get("answer_summary", ""), 80).replace("|", "\\|").replace("\n", " ")
        lines.append(
            f"| {i} | {rec.get('qa_id', '')[:12]} | {rec.get('domain', '')[:8]} | {q} | {a} | {ms} | _ | _ | _ | _ | _ | _ | _ | _ |"
        )

    output_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"  Generated: {output_path.name} ({len(records)} records, markdown fallback)")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Generate human review sheets")
    parser.add_argument("--input", type=str, default=str(INPUT_DEFAULT))
    parser.add_argument("--batch-size", type=int, default=20)
    parser.add_argument("--domain", type=str, default="")
    parser.add_argument("--format", choices=["xlsx", "md"], default="xlsx")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"ERROR: Input not found: {input_path}")
        print("Run auto_score.py first.")
        return

    records = []
    for line in input_path.read_text(encoding="utf-8").strip().split("\n"):
        if line.strip():
            try:
                rec = json.loads(line)
                if args.domain and rec.get("domain", "") != args.domain:
                    continue
                records.append(rec)
            except json.JSONDecodeError:
                continue

    if not records:
        print("No records to generate review sheets for.")
        return

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    num_batches = math.ceil(len(records) / args.batch_size)

    print(f"Generating {num_batches} review batch(es) from {len(records)} records...")

    for batch_idx in range(num_batches):
        start = batch_idx * args.batch_size
        end = min(start + args.batch_size, len(records))
        batch = records[start:end]

        suffix = ".xlsx" if args.format == "xlsx" and HAS_OPENPYXL else ".md"
        output_path = OUTPUT_DIR / f"review_batch_{batch_idx+1:03d}{suffix}"

        if args.format == "xlsx" and HAS_OPENPYXL:
            generate_xlsx(batch, output_path, batch_idx + 1)
        else:
            generate_markdown(batch, output_path, batch_idx + 1)

    print(f"\nAll review sheets saved to: {OUTPUT_DIR}/")
    if not HAS_OPENPYXL and args.format == "xlsx":
        print("NOTE: openpyxl not installed, generated markdown fallback.")
        print("  To get Excel: pip install openpyxl && python gen_review_sheet.py")


if __name__ == "__main__":
    main()
