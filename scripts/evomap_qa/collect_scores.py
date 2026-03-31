#!/usr/bin/env python3
"""collect_scores.py — 从 Excel/Markdown 评分表回收人工评分.

扫描 review_sheets/ 目录下的评分表，将人工评分写回 JSONL。

用法:
    python collect_scores.py                  # 扫描所有评分表
    python collect_scores.py --sheet path.xlsx # 指定单个文件

输出: 更新 planning_qa_scored.jsonl 中对应条目的 scores_human
"""
from __future__ import annotations

import argparse
import json
import re
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

SCORED_FILE = Path("/vol1/1000/projects/ChatgptREST/scripts/evomap_qa/planning_qa_scored.jsonl")
REVIEW_DIR = Path("/vol1/1000/projects/ChatgptREST/scripts/evomap_qa/review_sheets")

HUMAN_COLUMNS = ["clarity", "feasibility", "evidence", "risk", "alignment", "completeness", "overall"]
# Excel column indices (0-based from data area): cols 10-16 are the human scores, col 17 is comment
HUMAN_COL_START = 10  # Column J (1-indexed: 10)
COMMENT_COL = 17      # Column Q


def collect_from_xlsx(path: Path) -> dict[str, dict]:
    """Read human scores from an Excel file. Returns {qa_id: scores_dict}."""
    if not HAS_OPENPYXL:
        print(f"  SKIP (openpyxl not available): {path.name}")
        return {}

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    results: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=False):
        cells = list(row)
        if len(cells) < 17:
            continue

        qa_id = cells[1].value  # Column B
        if not qa_id:
            continue

        # Read human scores (columns J through P = indices 9-15)
        scores = {}
        has_any = False
        for i, dim in enumerate(HUMAN_COLUMNS):
            val = cells[HUMAN_COL_START - 1 + i].value
            if val is not None:
                try:
                    scores[dim] = float(val)
                    has_any = True
                except (ValueError, TypeError):
                    scores[dim] = None
            else:
                scores[dim] = None

        comment = cells[COMMENT_COL - 1].value if len(cells) >= COMMENT_COL else ""
        scores["comment"] = str(comment) if comment else ""

        if has_any:
            results[qa_id] = scores

    wb.close()
    return results


def collect_from_markdown(path: Path) -> dict[str, dict]:
    """Read human scores from a markdown table."""
    text = path.read_text(encoding="utf-8")

    results: dict[str, dict] = {}
    for line in text.split("\n"):
        if not line.startswith("|") or "---" in line:
            continue
        cells = [c.strip() for c in line.split("|")]
        cells = [c for c in cells if c]  # remove empty ends

        if len(cells) < 14:
            continue

        qa_id = cells[1]
        if not qa_id or qa_id == "QA ID" or qa_id == "#":
            continue

        # Cols: #, QA ID, 域, 问题, 答案, 机器分, 清晰度..完整度, 总分, 评语
        scores = {}
        has_any = False
        for i, dim in enumerate(HUMAN_COLUMNS):
            val = cells[6 + i] if 6 + i < len(cells) else "_"
            if val not in ("_", "", "-", "—"):
                try:
                    scores[dim] = float(val)
                    has_any = True
                except ValueError:
                    scores[dim] = None
            else:
                scores[dim] = None

        comment = cells[13] if len(cells) > 13 else ""
        scores["comment"] = comment if comment not in ("_", "-") else ""

        if has_any:
            results[qa_id] = scores

    return results


def update_jsonl(scored_map: dict[str, dict], scorer: str = "human") -> int:
    """Update the scored JSONL file with human scores."""
    if not SCORED_FILE.exists():
        print(f"ERROR: {SCORED_FILE} not found")
        return 0

    records = []
    for line in SCORED_FILE.read_text(encoding="utf-8").strip().split("\n"):
        if line.strip():
            try:
                records.append(json.loads(line))
            except json.JSONDecodeError:
                continue

    updated = 0
    for rec in records:
        qa_id = rec.get("qa_id", "")
        if qa_id in scored_map:
            human_scores = scored_map[qa_id]
            rec["scores_human"] = human_scores
            rec["human_scorer"] = scorer
            rec["human_scored_at"] = datetime.now(timezone.utc).isoformat()

            # Update status
            if human_scores.get("overall") is not None:
                rec["status"] = "scored"
            else:
                rec["status"] = "pending_human_review"

            # Check divergence
            rubric = rec.get("rubric_auto", {})
            if rubric and human_scores.get("overall") is not None:
                machine_avg = sum(rubric.values()) / max(len(rubric), 1)
                machine_5scale = 1 + machine_avg * 4
                human_overall = human_scores["overall"]
                divergence = abs(human_overall - machine_5scale)
                rec["divergence"] = round(divergence, 2)
                if divergence > 1.5:
                    rec["status"] = "needs_re_review"
                    rec["divergence_flag"] = True

            updated += 1

    # Write back
    with open(SCORED_FILE, "w", encoding="utf-8") as f:
        for rec in records:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")

    return updated


def main():
    parser = argparse.ArgumentParser(description="Collect human scores from review sheets")
    parser.add_argument("--sheet", type=str, default="",
                        help="Specific sheet file to process")
    parser.add_argument("--scorer", type=str, default="袁海州",
                        help="Name of human scorer")
    args = parser.parse_args()

    all_scores: dict[str, dict] = {}

    if args.sheet:
        sheets = [Path(args.sheet)]
    else:
        sheets = sorted(REVIEW_DIR.glob("review_batch_*"))

    for sheet in sheets:
        if not sheet.exists():
            print(f"SKIP (not found): {sheet}")
            continue

        print(f"Processing: {sheet.name}")
        if sheet.suffix == ".xlsx":
            scores = collect_from_xlsx(sheet)
        elif sheet.suffix == ".md":
            scores = collect_from_markdown(sheet)
        else:
            print(f"  SKIP (unknown format): {sheet.name}")
            continue

        filled = sum(1 for s in scores.values() if s.get("overall") is not None)
        print(f"  Found {len(scores)} entries, {filled} fully scored")
        all_scores.update(scores)

    if not all_scores:
        print("\nNo human scores found in review sheets.")
        print("Fill in the yellow columns in the Excel files first!")
        return

    updated = update_jsonl(all_scores, scorer=args.scorer)
    print(f"\n{'='*60}")
    print(f"Score Collection Complete")
    print(f"{'='*60}")
    print(f"Total entries with scores: {len(all_scores)}")
    print(f"Records updated in JSONL: {updated}")

    # Check divergence stats
    divergent = sum(1 for s in all_scores.values()
                    if s.get("overall") is not None)
    print(f"Fully scored: {divergent}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
